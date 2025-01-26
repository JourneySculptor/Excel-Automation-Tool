import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import gspread
import os
from fpdf import FPDF
from oauth2client.service_account import ServiceAccountCredentials

# Define the scope for Google Sheets API
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

CREDENTIALS_FILE = "credentials.json"

if not os.path.exists(CREDENTIALS_FILE):
    print(f"Error: Google credentials file '{CREDENTIALS_FILE}' not found. Check the file path.")
    exit()

creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, scope)

# User input
file_path = input("Enter the Excel file path (e.g., sample_data/sample.xlsx): ")

# Load data using pandas
try:
    df = pd.read_excel(file_path)
except FileNotFoundError:
    print("Error: File not found. Please check the file path and try again.")
    exit()
except Exception as e:
    print(f"Error loading Excel file: {e}")
    exit()

# Data Filtering
if "Sales" in df.columns:
    filtered_df = df[df["Sales"] > 10000]  # Filter sales above 10,000
    filtered_file = "sample_data/high_sales.xlsx"
    filtered_df.to_excel(filtered_file, index=False)
    print(f"Filtered data saved as: {filtered_file}")
else:
    print("Error: 'Sales' column not found in the dataset.")
    exit()

# Generate Sales Report Graph
if "Category" in df.columns and "Sales" in df.columns:
    category_sales = df.groupby("Category")["Sales"].sum()
    
    plt.figure(figsize=(8, 5))
    category_sales.plot(kind="bar", color='skyblue')
    plt.title("Sales by Category")
    plt.xlabel("Category")
    plt.ylabel("Total Sales")
    plt.xticks(rotation=0)
    graph_path = "sample_data/sales_chart.png"
    plt.savefig(graph_path)
    plt.show()
    print(f"Sales chart saved as: {graph_path}")
else:
    print("Error: 'Category' or 'Sales' column missing for graph.")
    exit()

# Generate PDF Report
try:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, "Sales Report", ln=True, align='C')
    pdf.ln(10)
    pdf.cell(200, 10, f"Filtered Data Saved: {filtered_file}", ln=True, align='L')
    pdf.cell(200, 10, f"Sales Chart Saved: {graph_path}", ln=True, align='L')

    pdf_file = "sample_data/sales_report.pdf"
    pdf.output(pdf_file)
    print(f"PDF Report generated: {pdf_file}")
except Exception as e:
    print(f"Error generating PDF: {e}")
    exit()

# Upload to Google Sheets
try:
    client = gspread.authorize(creds)
    
    sheet_name = "Excel_Automation_Report"
    spreadsheet = client.create(sheet_name)
    worksheet = spreadsheet.sheet1
    worksheet.update([filtered_df.columns.values.tolist()] + filtered_df.values.tolist())
    
    print(f"Data uploaded to Google Sheets: {sheet_name}")
except FileNotFoundError:
    print("Error: Google credentials file not found.")
except Exception as e:
    print(f"Google Sheets upload failed: {e}")
